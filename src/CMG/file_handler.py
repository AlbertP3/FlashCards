import openpyxl
import pandas as pd


class FileHandler:

    def __init__(self, path, sheet_name: str = None):
        self.path = path
        self.wb = openpyxl.load_workbook(self.path)
        self.ws = self.wb[sheet_name or self.wb.sheetnames[0]]

    @staticmethod
    def unshuffle_index(i: int, seed: int, datalen: int) -> int:
        si = (
            pd.DataFrame(data=[i for i in range(datalen)])
            .sample(frac=1, random_state=seed)
            .values.tolist()
        )
        return si[i][0]

    @staticmethod
    def unshuffle_dataframe(df: pd.DataFrame, seed: int) -> pd.DataFrame:
        si = [
            i[0]
            for i in pd.DataFrame(data=[i for i in range(df.shape[0])])
            .sample(frac=1, random_state=seed)
            .values.tolist()
        ]
        df = df.set_index(pd.Series(si), inplace=False).sort_index(inplace=False)
        return df

    def reverse_card(self, i: int, current_text: str, side: int) -> tuple[bool, str]:
        row = i + 2
        c = (self.ws.cell(row, 1).value, self.ws.cell(row, 2).value)
        if c[1 - side] != current_text:
            return False, "IndexValidationError: attempted edit on wrong cell!"
        self.ws.cell(row=row, column=1, value=c[1])
        self.ws.cell(row=row, column=2, value=c[0])
        self.wb.save(self.path)
        return True, "Successfully edited source file"

    def modify_card(self, i: int, new_card: list, old_card: list) -> tuple[bool, str]:
        row = i + 2
        c = [self.ws.cell(row, 1).value, self.ws.cell(row, 2).value]
        if c != old_card:
            return False, "IndexValidationError: attempted edit on wrong cell!"
        elif new_card == old_card:
            return False, "No need to change source file"
        self.ws.cell(row=row, column=1, value=new_card[1])
        self.ws.cell(row=row, column=2, value=new_card[0])
        self.wb.save(self.path)
        return True, "Successfully edited source file"

    def close(self):
        self.wb.close()
