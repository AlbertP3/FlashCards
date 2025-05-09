from abc import ABC, abstractmethod
import logging
import re
import pandas as pd
import openpyxl
from time import time
from collections import OrderedDict
from utils import get_filename_from_path
from cfg import config
from functools import cache
from csv import DictWriter

log = logging.getLogger("SOD")


class FileHandler(ABC):

    def __init__(self):
        self.path: str
        self.filename: str
        self.wb: openpyxl.Workbook
        self.native_lng: str
        self.foreign_lng: str
        self.raw_data: pd.DataFrame
        self.data: OrderedDict
        self.dtracker: tuple
        self.last_write_time: float

    @property
    @abstractmethod
    def total_rows(self):
        pass

    @abstractmethod
    def append_content(self, foreign_word: str, domestic_word: str) -> tuple[bool, str]:
        pass

    @abstractmethod
    def edit_content(self, foreign_word: str, domestic_word: str) -> tuple[bool, str]:
        pass

    @abstractmethod
    def delete_record(self, r0: str, r1: str, is_from_native: bool) -> bool:
        pass

    @abstractmethod
    def close(self):
        pass

    @staticmethod
    def reg_close(func):
        def inner(self, *args, **kwargs):
            path = self.path
            func(self, *args, **kwargs)
            log.debug(f"Closed FileHandler: {path}")

        return inner

    def get_languages(self) -> tuple:
        """returns values from header rows indicating languages used"""
        return self.native_lng, self.foreign_lng

    @cache
    def is_duplicate(self, word, is_from_native: bool) -> bool:
        """find duplicate, matching source_lng or native_lng"""
        i = int(is_from_native)
        for k, v in self.data.items():
            if word.lower() == v[i].lower():
                self.dtracker = (k, word)
                return True
        else:
            return False

    @cache
    def get_translations(self, phrase: str, is_from_native: bool) -> tuple:
        tran, orig = list(), list()
        i = int(is_from_native)
        for v in self.data.values():
            if phrase.lower() in v[i].lower():
                tran.append(v[1 - i])
                orig.append(v[i])
        return tran, orig

    @cache
    def get_translations_with_regex(
        self, phrase: str, is_from_native: bool
    ) -> tuple:
        pattern = re.compile(phrase, re.IGNORECASE)
        tran, orig = list(), list()
        i = int(is_from_native)
        for v in self.data.values():
            if pattern.search(v[i]):
                tran.append(v[1 - i])
                orig.append(v[i])
        return tran, orig

    def clear_cache(self):
        self.dtracker = None
        self.get_translations.cache_clear()
        self.is_duplicate.cache_clear()
        self.get_translations_with_regex.cache_clear()

    def update_dtracker(self, new_row: int = None, new_phrase: str = None):
        try:
            self.dtracker = (
                new_row or self.dtracker[0],
                new_phrase or self.dtracker[1],
            )
        except TypeError:
            return

    def validate_dtracker(self, searched_phrase: str) -> bool:
        """
        Returns True if dtracker is relevant to the current search,
        which also means it is a duplicate
        """
        try:
            return self.dtracker[1] == searched_phrase
        except TypeError:
            return False


class XLSXFileHandler(FileHandler):
    def __init__(self, path):
        self.path = path
        self.filename = get_filename_from_path(path, include_extension=True)
        self.wb = openpyxl.load_workbook(self.path, data_only=True)
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.__load_data()
        self.last_write_time = time()
        config["SOD"]["last_file"] = self.path
        self.native_lng = self.ws.cell(1, 2).value.lower()
        self.foreign_lng = self.ws.cell(1, 1).value.lower()

    def __load_data(self):
        self.data = OrderedDict()  # row_index: (original, translation)
        for r in range(1, self.ws.max_row + 1):
            self.data[r] = (
                str(self.ws.cell(r, 1).value),
                str(self.ws.cell(r, 2).value),
            )
        self.dtracker = None  # tuple: (row, searched_phrase)

    @property
    def total_rows(self):
        return self.ws.max_row - 1

    def append_content(self, foreign_word, domestic_word) -> tuple[bool, str]:
        target_row = self.ws.max_row + 1
        if self.ws.cell(row=target_row, column=1).value is None:
            self.ws.cell(row=target_row, column=1, value=foreign_word)
            self.ws.cell(row=target_row, column=2, value=domestic_word)
            self.data[target_row] = (foreign_word, domestic_word)
            self.commit(
                f"Added [{target_row}] [{foreign_word}] - [{domestic_word}] to {self.filename}"
            )
            return True, ""
        else:
            return False, "⚠ Target Cell is not empty!"

    def edit_content(self, foreign_word, domestic_word) -> tuple[bool, str]:
        self.ws.cell(row=self.dtracker[0], column=1, value=foreign_word)
        self.ws.cell(row=self.dtracker[0], column=2, value=domestic_word)
        self.data[self.dtracker[0]] = (foreign_word, domestic_word)
        self.commit(
            f"Edited [{self.dtracker[0]}] [{foreign_word}] - [{domestic_word}] in {self.filename}"
        )
        return True, ""

    def delete_record(self, r0: str, r1: str, is_from_native: bool) -> bool:
        if self.is_duplicate(r0, is_from_native):
            if (
                self.ws.cell(row=self.dtracker[0], column=int(is_from_native) + 1).value
                == r0
                and self.ws.cell(
                    row=self.dtracker[0], column=2 - int(is_from_native)
                ).value
                == r1
            ):
                self.ws.delete_rows(self.dtracker[0], 1)
                self.commit(f"Deleted [{self.dtracker[0]}] [{r0}] in {self.filename}")
                self.__load_data()
                return True
            else:
                return False
        else:
            return False

    def commit(self, msg: str = None):
        self.dtracker = None
        self.wb.save(self.path)
        self.last_write_time = time()
        self.clear_cache()
        if msg:
            log.info(msg)

    @FileHandler.reg_close
    def close(self):
        self.wb.close()
        del self


class CSVFileHandler(FileHandler):
    def __init__(self, path):
        self.path = path
        self.filename = get_filename_from_path(path, include_extension=True)
        self.__load_data()
        self.last_write_time = time()
        config["SOD"]["last_file"] = self.path
        self.native_lng = self.raw_data.columns[1].lower()
        self.foreign_lng = self.raw_data.columns[0].lower()

    def __load_data(self):
        self.raw_data = pd.read_csv(self.path, encoding="utf-8", dtype=str)
        self.data = OrderedDict()  # row_index: (original, translation)
        for r in self.raw_data.itertuples():
            self.data[r[0]] = (str(r[1]), str(r[2]))
        self.dtracker = None  # tuple: (row,)

    @property
    def total_rows(self):
        return self.raw_data.index[-1] + 1

    def append_content(self, foreign_word, domestic_word) -> tuple[bool, str]:
        new_row = (foreign_word, domestic_word)
        self.raw_data.loc[len(self.raw_data)] = new_row
        self.data[self.raw_data.index[-1]] = new_row
        self.__append(
            rec=new_row,
            msg=f"Added [{self.raw_data.index[-1]}] [{foreign_word}] - [{domestic_word}] to {self.filename}",
        )
        return True, ""

    def edit_content(self, foreign_word, domestic_word) -> tuple[bool, str]:
        edited_row = (foreign_word, domestic_word)
        self.data[self.dtracker[0]] = edited_row
        self.raw_data.iloc[self.dtracker[0]] = edited_row
        self.__edit(
            msg=f"Edited [{self.dtracker[0]}] [{foreign_word}] - [{domestic_word}] in {self.filename}",
        )
        return True, ""

    def delete_record(self, r0: str, r1: str, is_from_native: bool) -> bool:
        if self.is_duplicate(r0, is_from_native):
            if (
                self.raw_data.iloc[self.dtracker[0], int(is_from_native)] == r0
                and self.raw_data.iloc[self.dtracker[0], 1 - int(is_from_native)] == r1
            ):
                self.raw_data.drop(index=self.dtracker[0], axis=0, inplace=True)
                self.__edit(f"Deleted [{self.dtracker[0]}] [{r0}] in {self.filename}")
                self.__load_data()
                return True
            else:
                return False
        else:
            return False

    def __append(self, rec: tuple, msg: str = None):
        """Add a new record to the file"""
        with open(self.path, mode="a") as f:
            DictWriter(f, fieldnames=(self.foreign_lng, self.native_lng)).writerow(
                {self.foreign_lng: rec[0], self.native_lng: rec[1]}
            )
        self.last_write_time = time()
        self.clear_cache()
        if msg:
            log.info(msg)

    def __edit(self, msg: str = None):
        """Rewrites the file with modified content"""
        self.raw_data.to_csv(self.path, encoding="utf-8", index=False)
        self.last_write_time = time()
        self.clear_cache()
        if msg:
            log.info(msg)

    @FileHandler.reg_close
    def close(self):
        del self


class VoidFileHandler(FileHandler):
    def __init__(self, path=""):
        self.path = path
        self.filename = ""
        self.native_lng = ""
        self.foreign_lng = ""
        self.last_write_time = 0

    @property
    def total_rows(self):
        return 0

    def append_content(self, phrase, translations) -> tuple[bool, str]:
        return False, ""

    def edit_content(self, phrase, translations) -> tuple[bool, str]:
        return False, ""

    @FileHandler.reg_close
    def close(self):
        del self

    def is_duplicate(self, word, is_from_native):
        return False

    def get_translations(self, phrase: str, is_from_native: bool):
        return [], []

    def get_translations_with_regex(self, phrase: str, is_from_native: bool):
        return [], []

    def clear_cache(self):
        return

    def get_languages(self) -> tuple:
        return (self.foreign_lng, self.native_lng)

    def update_dtracker(self, new_row: int = None, new_phrase: str = None):
        return

    def validate_dtracker(self, searched_phrase: str):
        return

    def delete_record(self, r0: str, r1: str, is_from_native: bool) -> bool:
        return False

    def commit(self, msg: str = None):
        return


ACTIVE_FH: FileHandler = None


def get_filehandler(fullpath: str) -> FileHandler:
    """Pick FileHandler based on the file extension"""
    global ACTIVE_FH
    if fullpath.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        ACTIVE_FH = XLSXFileHandler(fullpath)
    elif fullpath.endswith(".csv"):
        ACTIVE_FH = CSVFileHandler(fullpath)
    elif fullpath.endswith(".void"):
        ACTIVE_FH = VoidFileHandler(fullpath)
    else:
        raise AttributeError(f'Unsupported file extension: {fullpath.split(".")[-1]}')
    log.debug(f"Initialized {type(ACTIVE_FH).__name__} for: {fullpath}")
    return ACTIVE_FH
