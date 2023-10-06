from abc import ABC, abstractmethod, abstractproperty
import re
import pandas as pd
import openpyxl
from collections import OrderedDict
from utils import Config, get_filename_from_path
from functools import cache

fhs = dict()



class FileHandler(ABC):

    @abstractproperty
    def total_rows(self):
        pass
    
    @abstractmethod
    def append_content(self) -> tuple[bool, str]:
        pass

    @abstractmethod
    def edit_content(self) -> tuple[bool, str]:
        pass

    @abstractmethod
    def close(self):
        pass

    def get_languages(self) -> tuple[str]:
        '''returns values from header rows indicating languages used'''
        return self.native_lng, self.foreign_lng
    
    @cache
    def is_duplicate(self, word, is_from_native:bool) -> bool:
        '''find duplicate, matching source_lng or native_lng'''
        if is_from_native:
            for k, r in self.data.items():
                if word == r[0]:
                    res, self.dtracker = True, (r[1], k)
                    break
            else:
                res = False
        else:
            try:
                res, self.dtracker = True, (self.data[word][1], word)
            except KeyError:
                res = False
        return res
    
    def get_translations(self, phrase:str, is_from_native:bool) -> tuple[list]:
        tran, orig = list(), list()
        if is_from_native:
            for k, v in self.data.items():
                if v[0] == phrase:
                    tran.append(k), orig.append(v[0])
        else:
            tran.append(self.data[phrase][0]), orig.append(phrase)
        return tran, orig

    def get_translations_with_regex(self, phrase:str, is_from_native:bool) -> tuple[list]:
        pattern = re.compile(phrase, re.IGNORECASE)
        tran, orig = list(), list()
        if is_from_native:
            for k, v in self.data.items():
                if pattern.search(v[0]):
                    tran.append(k), orig.append(v[0])
        else:
            for k, v in self.data.items():
                if pattern.search(k):
                    tran.append(v[0]), orig.append(k)
        return tran, orig



class XLSXFileHandler(FileHandler):

    def __init__(self, path):
        self.config = Config()
        self.path = path
        self.filename = get_filename_from_path(path, include_extension=True)
        self.wb = openpyxl.load_workbook(self.path)
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.data = OrderedDict()  # phrase: (translation, row_index)
        for r in range(1, self.ws.max_row+1):
            self.data[str(self.ws.cell(r, 1).value)] = (str(self.ws.cell(r, 2).value), r)
        self.config['SOD']['last_file'] = self.filename
        self.native_lng = self.ws.cell(1, 2).value.lower()
        self.foreign_lng = self.ws.cell(1, 1).value.lower()
        self.dtracker = None  # tuple: (row, phrase)
        fhs[self.path] = self


    @property
    def total_rows(self):
        return self.ws.max_row - 1
    

    def append_content(self, foreign_word, domestic_word) -> tuple[bool, str]:
        target_row = self.ws.max_row + 1
        if self.ws.cell(row=target_row, column=1).value is None:
            self.ws.cell(row=target_row, column=1, value=foreign_word)
            self.ws.cell(row=target_row, column=2, value=domestic_word)
            self.data[foreign_word] = (domestic_word, target_row)
            self.wb.save(self.path)
            self.is_duplicate.cache_clear()
            return True, ''
        else:
            return False, '[WARNING] Target Cell is not empty!'


    def edit_content(self, foreign_word, domestic_word) -> tuple[bool, str]:
        self.ws.cell(row=self.dtracker[0], column=1, value=foreign_word)
        self.ws.cell(row=self.dtracker[0], column=2, value=domestic_word)
        self.data[foreign_word] = (domestic_word, self.dtracker[0])
        if self.dtracker[1] != foreign_word:
            del self.data[self.dtracker[1]]
        self.dtracker = None
        self.wb.save(self.path)
        self.is_duplicate.cache_clear()
        return True, ''


    def close(self):
        self.wb.close()
        del fhs[self.path]



class CSVFileHandler(FileHandler):

    def __init__(self, path):
        self.config = Config()
        self.path = path
        self.filename = get_filename_from_path(path, include_extension=True)
        self.raw_data = pd.read_csv(self.path, encoding='utf-8')
        self.data = OrderedDict()  # phrase: (translation, row_index)
        for r in self.raw_data.itertuples():
            self.data[str(r[2])] = (str(r[1]), r[0])
        self.config['SOD']['last_file'] = self.filename
        self.native_lng = self.raw_data.columns[1].lower()
        self.foreign_lng = self.raw_data.columns[0].lower()
        self.dtracker = None  # tuple: (row, phrase)
        fhs[self.path] = self


    @property
    def total_rows(self):
        return self.raw_data.index[-1]
    

    def append_content(self, foreign_word, domestic_word) -> tuple[bool, str]:
        new_row = pd.DataFrame([[foreign_word, domestic_word]], columns=self.raw_data.columns)
        self.raw_data = pd.concat([self.raw_data, new_row], ignore_index=True)
        self.data[foreign_word] = (domestic_word, self.raw_data.index[-1])
        self.raw_data.to_csv(self.path, encoding='utf-8', index=False)
        return True, ''


    def edit_content(self, foreign_word, domestic_word) -> tuple[bool, str]:
        self.data[foreign_word] = (domestic_word, self.dtracker[0])
        self.raw_data.iloc[self.dtracker[0]] = [foreign_word, domestic_word]
        if self.dtracker[1] != foreign_word:
            del self.data[self.dtracker[1]]
        self.dtracker = None
        self.raw_data.to_csv(self.path, encoding='utf-8', index=False)
        self.is_duplicate.cache_clear()
        return True, ''


    def close(self):
        del fhs[self.path]



def get_filehandler(fullpath:str) -> FileHandler:
        '''Pick FileHandler based on the file extension'''
        if fullpath.endswith(('.xlsx','.xlsm','.xltx','.xltm')):
            return XLSXFileHandler(fullpath)
        elif fullpath.endswith('.csv'):
            return CSVFileHandler(fullpath)
        else:
            raise AttributeError(f'Unsupported file extension: {fullpath.split(".")[-1]}')